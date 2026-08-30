import json
import re
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Any
from uuid import NAMESPACE_URL, uuid5

from openpyxl import load_workbook
from sqlalchemy import delete, select
from sqlalchemy.orm import Session

from backend.infrastructure.models import (
    AuditEventRecord,
    CategoryRecord,
    FurnitureRecord,
    ImageRecord,
    InventoryAdjustmentRecord,
    InventoryRecord,
    SiteRecord,
)

MAIN_CATEGORY_EN = {
    "扶手椅和沙发": "Seating & Sofas",
    "储物家具": "Storage",
    "桌类": "Tables",
    "其他": "Other",
}

SITE_DIRECTORY = {
    "BJW": ("BJW", "北京西区", "北京", 116.31, 39.98),
    "Shang Hai": ("SHA", "上海园区", "上海", 121.47, 31.23),
    "Shanghai": ("SHA", "上海园区", "上海", 121.47, 31.23),
    "Shenzhen": ("SZX", "深圳园区", "深圳", 114.06, 22.54),
    "Su Zhou": ("SUZ", "苏州园区", "苏州", 120.58, 31.30),
    "Suzhou": ("SUZ", "苏州园区", "苏州", 120.58, 31.30),
    "Wu Xi": ("WUX", "无锡园区", "无锡", 120.31, 31.49),
    "Wuxi": ("WUX", "无锡园区", "无锡", 120.31, 31.49),
    "Guangzhou": ("CAN", "广州园区", "广州", 113.26, 23.13),
}

CATEGORY_TRANSLATIONS = {
    "工位椅-Zody": "Zody Task Chair",
    "会议椅-Very": "Very Conference Chair",
    "会议椅": "Conference Chair",
    "座椅，高脚凳/吧台凳": "Bar Stool",
    "座椅，高脚凳/ 吧台凳": "Bar Stool",
    "座椅，沙发/组合沙发": "Sofa / Sectional",
    "沙发组": "Sofa Set",
    "座椅，边椅/可折叠椅/餐椅": "Side / Folding / Dining Chair",
    "储物，盒子/文件抽屉柜": "Box / File Pedestal Cabinet",
    "储物，文件柜，对门": "Storage Cabinet",
    "储物，矮柜": "Low Storage Cabinet",
    "会议桌，圆形桌": "Round Meeting Table",
    "工位，升降桌": "Height-adjustable Workstation",
    "工位，办公桌": "Office Desk",
    "会议桌，D型桌": "D-shaped Meeting Table",
    "显示器支架（双臂/单臂）": "Monitor Arm (Dual / Single)",
    "演讲台": "Lectern",
}


@dataclass(frozen=True)
class ParsedFurniture:
    sku: str
    name: str
    name_en: str
    category: str
    category_en: str
    dimensions: str
    color: str
    material: str
    brand: str
    quantity: int
    image_url: str | None
    image_reference: str
    source_row: int
    source_metadata: dict[str, str | None]


@dataclass(frozen=True)
class ParsedSite:
    source_name: str
    code: str
    name: str
    city: str
    longitude: float
    latitude: float


@dataclass
class ImportReport:
    workbook: str
    dry_run: bool
    rows_seen: int = 0
    rows_imported: int = 0
    quantity_imported: int = 0
    categories_imported: int = 0
    sites_seen: int = 0
    sites_imported: int = 0
    images_embedded: int = 0
    images_recovered: int = 0
    image_reference_counts: dict[str, int] = field(default_factory=dict)
    warnings: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)

    def model_dump(self) -> dict[str, Any]:
        return asdict(self)


def clean(value: Any) -> str:
    return "" if value is None else str(value).strip()


def stable_id(kind: str, value: str) -> str:
    return str(uuid5(NAMESPACE_URL, f"furniture-center:{kind}:{value}"))


def parse_site(source_name: str) -> ParsedSite:
    if source_name not in SITE_DIRECTORY:
        raise ValueError(f"站点缺少可信地图坐标：{source_name}")
    code, name, city, longitude, latitude = SITE_DIRECTORY[source_name]
    return ParsedSite(source_name, code, name, city, longitude, latitude)


def bilingual_name(value: str) -> tuple[str, str]:
    compact = re.sub(r"\s+", " ", value.strip().rstrip(",")).strip()
    parts = [part.strip() for part in re.split(r"\s+/\s+", compact, maxsplit=1)]
    if len(parts) == 2 and re.search(r"[A-Za-z]", parts[1]):
        return parts[0], parts[1]
    return compact, CATEGORY_TRANSLATIONS.get(compact, "")


def image_status(value: Any) -> str:
    raw = clean(value)
    if raw == "#VALUE!":
        return "源文件图片引用失效"
    if raw == "有图片":
        return "源表标记有图片（媒体未随工作簿打包）"
    return raw


def parse_bjw(workbook_path: Path, report: ImportReport) -> list[ParsedFurniture]:
    workbook = load_workbook(workbook_path, data_only=False)
    if "BJW" not in workbook.sheetnames:
        report.errors.append("缺少 BJW 工作表")
        return []
    sheet = workbook["BJW"]
    report.images_embedded = len(sheet._images)
    headers = {clean(cell.value): cell.column for cell in sheet[2] if clean(cell.value)}
    required = {"家具类别（中/英）", "数量"}
    missing = sorted(required - headers.keys())
    if missing:
        report.errors.append(f"BJW 第 2 行缺少字段：{', '.join(missing)}")
        return []

    parsed: list[ParsedFurniture] = []
    main_category = "其他"
    for row in range(3, sheet.max_row + 1):
        raw_category = clean(sheet.cell(row, headers.get("类别", 1)).value)
        if raw_category:
            main_category = raw_category
        name_value = clean(sheet.cell(row, headers["家具类别（中/英）"]).value)
        quantity_value = sheet.cell(row, headers["数量"]).value
        if not name_value and quantity_value in (None, ""):
            continue
        report.rows_seen += 1
        try:
            quantity = int(quantity_value)
        except (TypeError, ValueError):
            report.errors.append(f"BJW 第 {row} 行数量无效：{quantity_value!r}")
            continue
        if quantity < 0:
            report.errors.append(f"BJW 第 {row} 行数量不能为负数")
            continue

        name, name_en = bilingual_name(name_value)
        raw_image = sheet.cell(row, headers.get("图片", 1)).value
        reference = image_status(raw_image)
        report.image_reference_counts[reference or "空白"] = (
            report.image_reference_counts.get(reference or "空白", 0) + 1
        )
        metadata = {
            header: clean(sheet.cell(row, column).value) or None
            for header, column in headers.items()
        }
        parsed.append(
            ParsedFurniture(
                sku=f"BJW-XLSX-{row:03d}",
                name=name,
                name_en=name_en,
                category=main_category,
                category_en=MAIN_CATEGORY_EN.get(main_category, ""),
                dimensions=clean(sheet.cell(row, headers.get("尺寸", 1)).value),
                color=clean(sheet.cell(row, headers.get("颜色", 1)).value),
                material=clean(sheet.cell(row, headers.get("材质", 1)).value),
                brand=clean(sheet.cell(row, headers.get("品牌", 1)).value),
                quantity=quantity,
                image_url=None,
                image_reference=reference,
                source_row=row,
                source_metadata=metadata,
            )
        )
    report.quantity_imported = sum(item.quantity for item in parsed)
    return parsed


def parse_manifest(
    manifest_path: Path, report: ImportReport
) -> tuple[list[ParsedFurniture], list[ParsedSite]]:
    try:
        payload = json.loads(manifest_path.read_text(encoding="utf-8-sig"))
    except (OSError, json.JSONDecodeError) as error:
        raise ValueError(f"无法读取迁移清单 '{manifest_path}': {error}") from error
    if payload.get("schema_version") != 1:
        raise ValueError("不支持的迁移清单版本")

    source_sheet = payload.get("source_sheet", "BJW")
    parsed: list[ParsedFurniture] = []
    sites: list[ParsedSite] = []
    for raw_site in payload.get("sites", [{"source_name": "BJW"}]):
        source_name = clean(raw_site.get("source_name"))
        if not source_name:
            continue
        report.sites_seen += 1
        try:
            site = parse_site(source_name)
        except ValueError as error:
            report.errors.append(str(error))
            continue
        if all(existing.code != site.code for existing in sites):
            sites.append(site)
    for raw in payload.get("rows", []):
        report.rows_seen += 1
        try:
            quantity = int(raw["quantity"])
            row = int(raw["row"])
        except (KeyError, TypeError, ValueError):
            report.errors.append(f"清单行数据无效：{raw!r}")
            continue
        name, name_en = bilingual_name(clean(raw.get("name")))
        image_url = clean(raw.get("image_url")) or None
        reference = clean(raw.get("image_reference"))
        if image_url:
            report.images_recovered += 1
        report.image_reference_counts[reference or "空白"] = (
            report.image_reference_counts.get(reference or "空白", 0) + 1
        )
        parsed.append(
            ParsedFurniture(
                sku=f"BJW-XLSX-{row:03d}",
                name=name,
                name_en=name_en,
                category=clean(raw.get("category")) or "其他",
                category_en=MAIN_CATEGORY_EN.get(clean(raw.get("category")), ""),
                dimensions=clean(raw.get("dimensions")),
                color=clean(raw.get("color")),
                material=clean(raw.get("material")),
                brand=clean(raw.get("brand")),
                quantity=quantity,
                image_url=image_url,
                image_reference=reference,
                source_row=row,
                source_metadata={
                    str(key): clean(value) or None
                    for key, value in dict(raw.get("metadata", {})).items()
                },
            )
        )
    report.quantity_imported = sum(item.quantity for item in parsed)
    report.images_embedded = report.images_recovered
    report.workbook = clean(payload.get("source_workbook")) or manifest_path.name
    if source_sheet != "BJW":
        report.warnings.append(f"清单来源工作表为 {source_sheet}，按 BJW 格式导入。")
    return parsed, sites


def import_workbook(
    session: Session,
    workbook_path: Path,
    *,
    dry_run: bool = True,
    replace_catalog: bool = False,
) -> ImportReport:
    if not workbook_path.exists():
        raise ValueError(f"找不到工作簿：{workbook_path}")
    report = ImportReport(workbook=workbook_path.name, dry_run=dry_run)
    if workbook_path.suffix.lower() == ".json":
        parsed, parsed_sites = parse_manifest(workbook_path, report)
    else:
        parsed = parse_bjw(workbook_path, report)
        parsed_sites = [parse_site("BJW")]
        report.sites_seen = 1
    if report.images_embedded == 0:
        report.warnings.append("工作簿未包含可提取的嵌入图片；仅保留图片列状态。")
    if report.errors or dry_run:
        return report

    if replace_catalog:
        session.execute(delete(InventoryAdjustmentRecord))
        session.execute(delete(InventoryRecord))
        session.execute(delete(ImageRecord))
        session.execute(delete(FurnitureRecord))
        session.execute(delete(AuditEventRecord))
        session.execute(delete(CategoryRecord))
        session.execute(delete(SiteRecord))
        session.flush()

    sites: dict[str, SiteRecord] = {}
    for parsed_site in parsed_sites:
        site_id = stable_id("site", parsed_site.code)
        site = session.get(SiteRecord, site_id)
        if site is None:
            site = SiteRecord(id=site_id, code=parsed_site.code)
            session.add(site)
        site.name = parsed_site.name
        site.city = parsed_site.city
        site.longitude = parsed_site.longitude
        site.latitude = parsed_site.latitude
        sites[parsed_site.code] = site
        report.sites_imported += 1
    if "BJW" not in sites:
        raise ValueError("迁移清单必须包含 BJW 站点")
    site = sites["BJW"]
    site_id = site.id

    categories: dict[str, CategoryRecord] = {}
    for name in sorted({item.category for item in parsed}):
        category_id = stable_id("category", name)
        category = session.get(CategoryRecord, category_id)
        if category is None:
            category = CategoryRecord(id=category_id, name=name)
            session.add(category)
        categories[name] = category
    report.categories_imported = len(categories)
    session.flush()

    for item in parsed:
        furniture_id = stable_id("furniture", item.sku)
        record = session.scalar(select(FurnitureRecord).where(FurnitureRecord.sku == item.sku))
        if record is None:
            record = FurnitureRecord(id=furniture_id, sku=item.sku)
            session.add(record)
        record.name = item.name
        record.name_en = item.name_en
        record.main_category = item.category
        record.category = categories[item.category]
        record.description = item.name_en or f"{item.name}，来源于 BJW 复用家具台账。"
        record.condition = "good"
        record.dimensions = item.dimensions
        record.color = item.color
        record.material = item.material
        record.brand = item.brand
        record.image_reference = item.image_reference
        record.source_workbook = workbook_path.name
        record.source_sheet = "BJW"
        record.source_row = item.source_row
        record.source_metadata = json.dumps(item.source_metadata, ensure_ascii=False)

        if item.image_url:
            primary_image = next((image for image in record.images if image.is_primary), None)
            if primary_image is None:
                primary_image = ImageRecord(
                    id=stable_id("image", item.sku),
                    furniture=record,
                    url=item.image_url,
                    alt_text=item.name,
                    is_primary=True,
                )
                session.add(primary_image)
            else:
                primary_image.url = item.image_url
                primary_image.alt_text = item.name

        position = next((entry for entry in record.inventory if entry.site_id == site_id), None)
        if position is None:
            position = InventoryRecord(
                id=stable_id("inventory", f"{item.sku}:BJW"),
                furniture=record,
                site=site,
                quantity_total=item.quantity,
                quantity_available=item.quantity,
            )
            session.add(position)
        else:
            position.quantity_total = item.quantity
            position.quantity_available = item.quantity
        report.rows_imported += 1

    session.commit()
    return report